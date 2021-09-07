from openpyxl import load_workbook #.xlsx modification
import pandas as pd #.csv -> .xlsx conversion

read_file = pd.read_csv(r'./data.csv', sep = '\t')
saveFile = pd.ExcelWriter('data.xlsx')
read_file.to_excel(saveFile, index = None, header=True)

saveFile.save()

f = open('genes.txt', 'r')
genes = f.read()
f.close()

genes = genes.split(" ")

wb = load_workbook('data.xlsx')

ws = wb.active


rows = len(tuple(ws.rows))
cols = len(tuple(ws.columns))

for i in range(rows):
    for j in range(3):
        if i != 0:
            if j == 0:
                word = 'prot'
            elif j == 1:
                word = 'gseq'
            else:
                word = 'gint'
            ws.cell(row=i+1,column=(cols-j),value='=Hyperlink("{}","{}")'.format(genes[i-1]+"/"+word+genes[i-1]+".fasta", word+genes[i-1]))

wb.save('data.xlsx')

#check to see if the excel file has been run through cleanly
#i added this here since from main.py it runs os.system('python excelMaker')
#and it doesnt output an error if anythings wrong
print("excelMaker finished running")